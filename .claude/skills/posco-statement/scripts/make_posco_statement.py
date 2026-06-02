#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create POSCO E&C transaction statement PDF/workbook from eCount sales data."""

from __future__ import annotations

import argparse
import copy
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import MergedCell

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except ModuleNotFoundError:
    REPORTLAB_AVAILABLE = False


CODE_RE = re.compile(r"(?<![A-Z0-9])A\s*0*(\d{1,7})(?:\s*\*\s*(-?\d+(?:\.\d+)?))?", re.I)
INVALID_FILENAME_RE = re.compile(r'[<>:"/\\|?*\x00-\x1F]')
RAW_REQUIRED_HEADERS = {"일자-No.", "거래처명", "프로젝트명", "품목명", "수량", "공급가액", "적요"}


@dataclass
class CodeInfo:
    code: str
    item: str
    unit: str
    price: Decimal


@dataclass
class StatementLine:
    source_row: int
    date_no: str
    date: datetime | None
    project: str
    vendor: str
    raw_item: str
    raw_spec: str
    raw_qty: Decimal
    raw_supply: Decimal | None
    token: str
    buyer_code: str
    item: str
    unit: str
    unit_price: Decimal
    qty: Decimal
    amount: Decimal
    note: str
    remark: str


def clean_text(value: Any) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def sanitize_filename(name: str, fallback: str = "포스코_거래명세서") -> str:
    value = unicodedata.normalize("NFC", clean_text(name))
    value = INVALID_FILENAME_RE.sub("_", value)
    value = re.sub(r"\s+", " ", value).strip(" .")
    return value or fallback


def safe_sheet_title(name: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "-", clean_text(name))
    base = re.sub(r"\s+", " ", base).strip() or "명세서"
    base = base[:31]
    title = base
    i = 2
    while title in used:
        suffix = f" {i}"
        title = (base[:31 - len(suffix)] + suffix).strip()
        i += 1
    used.add(title)
    return title


def decimal_value(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value is None or value == "":
        return default
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError, AttributeError):
        return default


def money(value: Decimal) -> str:
    rounded = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if rounded == rounded.to_integral():
        return f"{int(rounded):,}"
    return f"{rounded:,.2f}"


def qty_text(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return f"{int(normalized):,}"
    return f"{normalized:f}".rstrip("0").rstrip(".")


def parse_date_no(value: Any) -> tuple[str, datetime | None, str, str]:
    s = clean_text(value)
    m = re.search(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})\s*-\s*(\d+)", s)
    if not m:
        return s, None, "", ""
    dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return s, dt, str(int(m.group(2))), str(int(m.group(3)))


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        return clean_text(value.get("text") or value.get("result") or "")
    return clean_text(value)


def find_raw_sheet_and_headers(wb):
    for ws in wb.worksheets:
        max_r = min(ws.max_row, 25)
        max_c = min(ws.max_column, 60)
        for r in range(1, max_r + 1):
            values = [cell_text(ws.cell(r, c).value) for c in range(1, max_c + 1)]
            found = {v for v in values if v}
            if len(RAW_REQUIRED_HEADERS.intersection(found)) >= 6:
                header_map = {}
                for i, value in enumerate(values):
                    if value and value not in header_map:
                        header_map[value] = i
                return ws, r, header_map
    raise RuntimeError("판매현황 헤더를 찾지 못했습니다. 원본 eCount 판매현황 파일인지 확인해주세요.")


def get_row_value(row: tuple[Any, ...], header_map: dict[str, int], name: str) -> Any:
    idx = header_map.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def find_code_sheet(wb):
    for ws in wb.worksheets:
        name = clean_text(ws.title)
        if "단가계약" in name or "품목" in name:
            return ws
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 20) + 1):
            row_text = " ".join(clean_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 15) + 1))
            if "25년 코드" in row_text and "품명" in row_text:
                return ws
    raise RuntimeError("포스코 단가계약 품목 코드표 시트를 찾지 못했습니다.")


def build_code_index(template_path: str) -> dict[str, CodeInfo]:
    wb = load_workbook(template_path, data_only=True)
    ws = find_code_sheet(wb)
    index: dict[str, CodeInfo] = {}
    for r in range(1, ws.max_row + 1):
        code = clean_text(ws.cell(r, 2).value).upper().replace(" ", "")
        if not re.fullmatch(r"A\d{7}", code):
            continue
        item = clean_text(ws.cell(r, 4).value)
        unit = clean_text(ws.cell(r, 5).value)
        price = decimal_value(ws.cell(r, 6).value)
        info = CodeInfo(code=code, item=item, unit=unit, price=price)
        digits = code[1:]
        index[code] = info
        index[digits] = info
        if code.startswith("A2500"):
            short = code[5:]
            index[short] = info
            index[str(int(short))] = info
            index[f"A{short}"] = info
            index[f"A{int(short):03d}"] = info
    return index


def lookup_code(code_index: dict[str, CodeInfo], digits: str) -> CodeInfo | None:
    raw = re.sub(r"\D", "", digits)
    if not raw:
        return None
    keys = [raw, f"A{raw}"]
    if len(raw) <= 4:
        n = int(raw)
        keys.extend([str(n), f"{n:03d}", f"A{n:03d}", f"A2500{n:03d}"])
    if raw.startswith("2500"):
        keys.extend([f"A{raw}", raw[-3:], str(int(raw[-3:])), f"A{raw[-3:]}"])
    for key in keys:
        found = code_index.get(key.upper() if key.startswith("A") else key)
        if found:
            return found
    return None


def remark_text(row: tuple[Any, ...], header_map: dict[str, int]) -> str:
    return " ".join(
        clean_text(get_row_value(row, header_map, name))
        for name in ("적요", "적요2", "적요3")
        if clean_text(get_row_value(row, header_map, name))
    )


def note_for_match(text: str, match: re.Match) -> str:
    after = text[match.end():]
    if "/" in after:
        return re.sub(r"\s+", " ", after.split("/", 1)[1]).strip(" /|")
    after = re.sub(r"^[\s+|,/:-]+", "", after)
    after = re.sub(r"\s+", " ", after).strip()
    return after


def read_statement_lines(raw_path: str, template_path: str, include_mismatch: bool = False) -> tuple[list[StatementLine], list[str]]:
    code_index = build_code_index(template_path)
    wb = load_workbook(raw_path, data_only=True)
    ws, header_row, header_map = find_raw_sheet_and_headers(wb)
    lines: list[StatementLine] = []
    warnings: list[str] = []
    row_amounts: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    row_supply: dict[int, Decimal | None] = {}
    row_labels: dict[int, str] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        text = remark_text(row, header_map)
        if not text or not CODE_RE.search(text):
            continue
        date_no, dt, _, _ = parse_date_no(get_row_value(row, header_map, "일자-No."))
        project = clean_text(get_row_value(row, header_map, "프로젝트명"))
        vendor = clean_text(get_row_value(row, header_map, "거래처명"))
        raw_item = clean_text(get_row_value(row, header_map, "품목명"))
        raw_spec = clean_text(get_row_value(row, header_map, "규격"))
        raw_qty = decimal_value(get_row_value(row, header_map, "수량"))
        raw_supply_val = get_row_value(row, header_map, "공급가액")
        raw_supply = None if raw_supply_val in (None, "") else decimal_value(raw_supply_val)
        row_supply[row_idx] = raw_supply
        row_labels[row_idx] = f"{date_no} / {raw_item} / {text}"

        for match in CODE_RE.finditer(text):
            info = lookup_code(code_index, match.group(1))
            token = re.sub(r"\s+", "", match.group(0).upper())
            if not info:
                warnings.append(f"row {row_idx}: 코드 {token}을 템플릿 코드표에서 찾지 못해 제외했습니다.")
                continue
            qty = decimal_value(match.group(2)) if match.group(2) is not None else raw_qty
            amount = info.price * qty
            note = note_for_match(text, match)
            if not note:
                note = " / ".join(x for x in [raw_item, raw_spec] if x)
            line = StatementLine(
                source_row=row_idx,
                date_no=date_no,
                date=dt,
                project=project,
                vendor=vendor,
                raw_item=raw_item,
                raw_spec=raw_spec,
                raw_qty=raw_qty,
                raw_supply=raw_supply,
                token=token,
                buyer_code=info.code,
                item=info.item,
                unit=info.unit,
                unit_price=info.price,
                qty=qty,
                amount=amount,
                note=note,
                remark=text,
            )
            lines.append(line)
            row_amounts[row_idx] += amount

    mismatch_rows: set[int] = set()
    for row_idx, supply in row_supply.items():
        if supply is None or row_amounts[row_idx] == 0:
            continue
        diff = (row_amounts[row_idx] - supply).copy_abs()
        if diff > Decimal("2"):
            mismatch_rows.add(row_idx)
            action = "포함" if include_mismatch else "보류"
            warnings.append(
                f"{action}: row {row_idx}: {row_labels.get(row_idx, '')} / 코드 계산금액 {money(row_amounts[row_idx])} / 원본 공급가액 {money(supply)} 차이 {money(diff)}"
            )

    if mismatch_rows and not include_mismatch:
        lines = [line for line in lines if line.source_row not in mismatch_rows]

    if not lines:
        raise RuntimeError("적요에서 A283 같은 포스코 코드가 있는 행을 찾지 못했습니다.")
    return lines, warnings


def find_statement_sheet(wb):
    for ws in wb.worksheets:
        if "거래명세" in clean_text(ws.title) or "명세서" in clean_text(ws.title):
            return ws
    for ws in wb.worksheets:
        if clean_text(ws.cell(1, 1).value) == "거래 명세표":
            return ws
    raise RuntimeError("거래명세서 템플릿 시트를 찾지 못했습니다.")


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def clear_statement_rows(ws) -> None:
    for r in range(8, 31):
        for c in (1, 2, 3, 9, 11):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None


def set_formula_recalc(wb) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def populate_workbook(template_path: str, output_path: str, groups: list[tuple[str, list[StatementLine]]]) -> None:
    wb = load_workbook(template_path)
    template_ws = find_statement_sheet(wb)
    used = set(wb.sheetnames)
    created = []

    for group_key, group_lines in groups:
        chunks = [group_lines[i:i + 23] for i in range(0, len(group_lines), 23)]
        for chunk_idx, chunk in enumerate(chunks, start=1):
            ws = wb.copy_worksheet(template_ws)
            suffix = f"-{chunk_idx}" if len(chunks) > 1 else ""
            ws.title = safe_sheet_title(f"{group_key}{suffix}", used)
            clear_statement_rows(ws)
            created.append(ws.title)
            for offset, line in enumerate(chunk, start=8):
                month = line.date.month if line.date else ""
                day = line.date.day if line.date else ""
                ws.cell(offset, 1).value = month
                ws.cell(offset, 2).value = day
                ws.cell(offset, 3).value = line.buyer_code
                ws.cell(offset, 9).value = float(line.qty)
                ws.cell(offset, 11).value = line.note

    if template_ws.title in wb.sheetnames and len(created) > 0:
        try:
            template_ws.sheet_state = "hidden"
        except Exception:
            pass
    set_formula_recalc(wb)
    wb.save(output_path)


def register_fonts() -> tuple[str, str]:
    windir = os.environ.get("WINDIR", r"C:\Windows")
    candidates = [
        (os.path.join(windir, "Fonts", "malgun.ttf"), os.path.join(windir, "Fonts", "malgunbd.ttf")),
        (r"C:\Windows\Fonts\malgun.ttf", r"C:\Windows\Fonts\malgunbd.ttf"),
    ]
    for regular, bold in candidates:
        if os.path.exists(regular):
            pdfmetrics.registerFont(TTFont("Korean", regular))
            if os.path.exists(bold):
                pdfmetrics.registerFont(TTFont("KoreanBold", bold))
            else:
                pdfmetrics.registerFont(TTFont("KoreanBold", regular))
            return "Korean", "KoreanBold"
    return "Helvetica", "Helvetica-Bold"


def paragraph(text: Any, style: ParagraphStyle) -> Paragraph:
    escaped = (
        clean_text(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(escaped, style)


def pdf_payload(groups: list[tuple[str, list[StatementLine]]], warnings: list[str]) -> dict[str, Any]:
    out_groups = []
    for group_key, lines in groups:
        first = lines[0]
        supply = sum((line.amount for line in lines), Decimal("0"))
        tax = (supply * Decimal("0.1")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total = supply + tax
        out_groups.append({
            "key": group_key,
            "project": first.project,
            "issueDate": datetime.now().strftime("%Y-%m-%d"),
            "supplyText": money(supply),
            "taxText": money(tax),
            "totalText": money(total),
            "lines": [{
                "month": str(line.date.month) if line.date else "",
                "day": str(line.date.day) if line.date else "",
                "buyerCode": line.buyer_code,
                "item": line.item,
                "unitPrice": money(line.unit_price),
                "unit": line.unit,
                "qty": qty_text(line.qty),
                "amount": money(line.amount),
                "note": line.note,
            } for line in lines],
        })
    return {"groups": out_groups, "warnings": warnings}


def draw_statement_pdf_node(output_path: str, groups: list[tuple[str, list[StatementLine]]], warnings: list[str]) -> None:
    node = os.environ.get("NODE_EXE") or shutil.which("node")
    if not node:
        raise RuntimeError("PDF 생성을 위해 reportlab 또는 Node.js가 필요합니다. 서버에 node 또는 reportlab을 확인해주세요.")
    script = os.path.join(os.path.dirname(__file__), "render_posco_pdf.js")
    if not os.path.exists(script):
        raise RuntimeError(f"PDF 렌더러를 찾지 못했습니다: {script}")
    fd, payload_path = tempfile.mkstemp(suffix=".json")
    os.close(fd)
    try:
        output_abs = os.path.abspath(output_path)
        with open(payload_path, "w", encoding="utf-8") as f:
            json.dump(pdf_payload(groups, warnings), f, ensure_ascii=False)
        result = subprocess.run(
            [node, script, payload_path, output_abs],
            cwd=os.path.dirname(script),
            text=True,
            encoding="utf-8",
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120,
        )
        if result.returncode != 0:
            tail = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(f"Node PDF 렌더링 실패: {tail}")
    finally:
        try:
            os.remove(payload_path)
        except OSError:
            pass


def draw_statement_pdf(output_path: str, groups: list[tuple[str, list[StatementLine]]], warnings: list[str]) -> None:
    if os.environ.get("POSCO_FORCE_NODE_PDF") == "1" or not REPORTLAB_AVAILABLE:
        return draw_statement_pdf_node(output_path, groups, warnings)

    font, bold_font = register_fonts()
    styles = {
        "title": ParagraphStyle("title", fontName=bold_font, fontSize=20, leading=24, alignment=TA_CENTER),
        "normal": ParagraphStyle("normal", fontName=font, fontSize=8, leading=10, alignment=TA_LEFT),
        "small": ParagraphStyle("small", fontName=font, fontSize=7, leading=9, alignment=TA_LEFT),
        "right": ParagraphStyle("right", fontName=font, fontSize=8, leading=10, alignment=TA_RIGHT),
        "center": ParagraphStyle("center", fontName=font, fontSize=8, leading=10, alignment=TA_CENTER),
        "bold": ParagraphStyle("bold", fontName=bold_font, fontSize=9, leading=11, alignment=TA_LEFT),
    }
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    story = []

    for group_idx, (group_key, lines) in enumerate(groups):
        first = lines[0]
        supply = sum((line.amount for line in lines), Decimal("0"))
        tax = (supply * Decimal("0.1")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total = supply + tax
        issue_date = datetime.now().strftime("%Y-%m-%d")

        if group_idx:
            story.append(PageBreak())
        story.append(paragraph("거래 명세표", styles["title"]))
        story.append(Spacer(1, 4))

        header = Table([
            [
                paragraph(f"작성일: {issue_date}", styles["normal"]),
                paragraph(f"일자-No.: {group_key}", styles["right"]),
            ],
            [
                paragraph(f"공급받는자: ㈜ 포스코이앤씨\n현장: {first.project}", styles["normal"]),
                paragraph("공급자: (주)엔투비\n등록번호: 220-81-96244\n서울시 강남구 봉은사로 514 포스코타워-삼성", styles["normal"]),
            ],
            [
                paragraph(f"합계 금액: {money(total)} 원정", styles["bold"]),
                paragraph(f"공급가액 {money(supply)} / 부가세 {money(tax)}", styles["right"]),
            ],
        ], colWidths=[93 * mm, 93 * mm])
        header.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#f2f2f2")),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(header)
        story.append(Spacer(1, 6))

        table_data = [[
            paragraph("월", styles["center"]),
            paragraph("일", styles["center"]),
            paragraph("구매사코드", styles["center"]),
            paragraph("품목", styles["center"]),
            paragraph("단가", styles["center"]),
            paragraph("단위", styles["center"]),
            paragraph("수량", styles["center"]),
            paragraph("금액", styles["center"]),
            paragraph("비고", styles["center"]),
        ]]
        for line in lines:
            table_data.append([
                paragraph(str(line.date.month) if line.date else "", styles["center"]),
                paragraph(str(line.date.day) if line.date else "", styles["center"]),
                paragraph(line.buyer_code, styles["center"]),
                paragraph(line.item, styles["small"]),
                paragraph(money(line.unit_price), styles["right"]),
                paragraph(line.unit, styles["center"]),
                paragraph(qty_text(line.qty), styles["right"]),
                paragraph(money(line.amount), styles["right"]),
                paragraph(line.note, styles["small"]),
            ])
        table = Table(table_data, colWidths=[8 * mm, 8 * mm, 23 * mm, 50 * mm, 20 * mm, 12 * mm, 16 * mm, 24 * mm, 25 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eeeeee")),
            ("FONTNAME", (0, 0), (-1, 0), bold_font),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
        story.append(Spacer(1, 6))
        totals = Table([
            ["", "", paragraph("공급가액", styles["center"]), paragraph(money(supply), styles["right"])],
            ["", "", paragraph("부가세", styles["center"]), paragraph(money(tax), styles["right"])],
            ["", "", paragraph("합계", styles["center"]), paragraph(money(total), styles["right"])],
        ], colWidths=[90 * mm, 40 * mm, 25 * mm, 31 * mm])
        totals.setStyle(TableStyle([
            ("GRID", (2, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (2, 2), (-1, 2), colors.HexColor("#f2f2f2")),
            ("FONTNAME", (2, 2), (-1, 2), bold_font),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(totals)

    if warnings:
        story.append(PageBreak())
        story.append(paragraph("확인 필요 항목", styles["title"]))
        story.append(Spacer(1, 6))
        for warning in warnings[:80]:
            story.append(paragraph("- " + warning, styles["normal"]))
            story.append(Spacer(1, 2))

    doc.build(story)


def grouped_lines(lines: list[StatementLine]) -> list[tuple[str, list[StatementLine]]]:
    buckets: dict[str, list[StatementLine]] = defaultdict(list)
    for line in lines:
        buckets[line.date_no].append(line)
    def sort_key(item):
        key, group = item
        first = group[0]
        return (first.date or datetime.min, key, min(x.source_row for x in group))
    groups = sorted(buckets.items(), key=sort_key)
    for _, group in groups:
        group.sort(key=lambda x: (x.source_row, x.buyer_code))
    return groups


def make_outputs(raw_path: str, template_path: str, outdir: str, include_mismatch: bool = False) -> dict[str, Any]:
    os.makedirs(outdir, exist_ok=True)
    lines, warnings = read_statement_lines(raw_path, template_path, include_mismatch=include_mismatch)
    groups = grouped_lines(lines)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"포스코_거래명세서_{stamp}"
    xlsx_path = os.path.join(outdir, base + ".xlsx")
    pdf_path = os.path.join(outdir, base + ".pdf")
    populate_workbook(template_path, xlsx_path, groups)
    draw_statement_pdf(pdf_path, groups, warnings)
    return {
        "pdf": pdf_path,
        "xlsx": xlsx_path,
        "groups": len(groups),
        "lines": len(lines),
        "warnings": warnings,
    }


def app_root() -> str:
    p = os.path.abspath(__file__)
    for _ in range(8):
        p = os.path.dirname(p)
        if os.path.exists(os.path.join(p, "server.js")):
            return p
    return ""


def valid_posco_template(path: str) -> bool:
    try:
        if not path or not os.path.isfile(path):
            return False
        name = os.path.basename(path)
        if name.startswith("~$") or not name.lower().endswith(".xlsx"):
            return False
        build_code_index(path)
        wb = load_workbook(path, data_only=False)
        find_statement_sheet(wb)
        return True
    except Exception:
        return False


def resolve_template_arg(template_arg: str) -> str:
    candidates = []
    if template_arg:
        if os.path.isfile(template_arg) and valid_posco_template(template_arg):
            return template_arg
        if os.path.isdir(template_arg):
            candidates.extend(str(p) for p in Path(template_arg).glob("*.xlsx"))
    root = app_root()
    if root:
        candidates.extend(str(p) for p in Path(root, "data", "ai-skill-templates", "posco-statement").glob("*.xlsx"))
        candidates.extend(str(p) for p in Path(root, "learning-data", "_무관_포스코").glob("*.xlsx"))
    valid = [p for p in candidates if valid_posco_template(p)]
    if not valid:
        return ""
    return max(valid, key=lambda p: (os.path.getsize(p), os.path.getmtime(p)))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Create POSCO transaction statement PDF/workbook.")
    parser.add_argument("--raw", required=True, help="eCount 판매현황 xlsx")
    parser.add_argument("--template", default="", help="POSCO 거래명세서 tool/template xlsx or template directory")
    parser.add_argument("--outdir", default=".", help="output directory")
    parser.add_argument("--include-mismatch", action="store_true", help="include rows whose generated amount does not match raw 공급가액")
    args = parser.parse_args(argv)

    try:
        template = resolve_template_arg(args.template)
        if not template:
            raise RuntimeError("POSCO 거래명세서 툴 템플릿이 없습니다. data/ai-skill-templates/posco-statement 폴더에 템플릿 xlsx를 등록해야 합니다.")
        print(f"템플릿: {template}")
        result = make_outputs(args.raw, template, args.outdir, include_mismatch=args.include_mismatch)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"POSCO 거래명세서 생성 완료: {result['groups']}개 명세서 / {result['lines']}개 항목")
    print(f"PDF: {result['pdf']}")
    print(f"XLSX: {result['xlsx']}")
    if result["warnings"]:
        print("확인 필요:")
        for w in result["warnings"][:30]:
            print(f"- {w}")
        if len(result["warnings"]) > 30:
            print(f"- ... 외 {len(result['warnings']) - 30}건")
    print(json.dumps({
        "ok": True,
        "pdf": result["pdf"],
        "xlsx": result["xlsx"],
        "groups": result["groups"],
        "lines": result["lines"],
        "warnings": result["warnings"],
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
