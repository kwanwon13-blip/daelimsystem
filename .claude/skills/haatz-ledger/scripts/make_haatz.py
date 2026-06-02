#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
하츠 마감내역서(거래명세서) 자동 생성 — ERP 에이전트용

나이스텍과 같은 단일 '안전' 시트 / 부가세별도 양식. 단, 하츠는:
  - 현장(프로젝트명)별로 파일 분리
  - E4 = '거래처: {현장명}' 으로 설정(덮어씀) — 나이스텍과 달리 현장명을 넣는다
  - 파일명: '{YY}년 {MM}월마감 하츠 마감내역서 - {현장}.xlsx'

사용:
  python make_haatz.py --raw <판매현황.xlsx> --template <전월 마감내역서 폴더/파일> --outdir <폴더>
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import copy, shutil, os, unicodedata, re, tempfile, argparse, sys
from datetime import datetime

SKIP_ITEMS = {'매출할인'}

def parse_date(s):
    m=re.match(r'(\d{4}/\d{2}/\d{2})', str(s))
    return datetime.strptime(m.group(1),'%Y/%m/%d') if m else None

def copy_style(src,dst):
    if src.has_style:
        dst.font=copy.copy(src.font); dst.border=copy.copy(src.border)
        dst.fill=copy.copy(src.fill); dst.number_format=src.number_format
        dst.alignment=copy.copy(src.alignment)

def find_summary_rows(ws):
    s=v=t=None
    for r in ws.iter_rows():
        for cell in r:
            if isinstance(cell, MergedCell): continue
            val=str(cell.value).strip() if cell.value else ''
            if val=='공급금액': s=cell.row
            elif val=='부가세': v=cell.row
            elif val=='총합계': t=cell.row
    return s,v,t

def fill_sheet(ws, rows, data_start=13):
    supply_row, vat_row, total_row = find_summary_rows(ws)
    if supply_row is None:
        raise RuntimeError("템플릿에서 '공급금액' 합계 행을 못 찾음")
    available=supply_row-data_start; needed=len(rows)
    if needed>available:
        add=needed-available
        ref={c:ws.cell(row=data_start,column=c) for c in range(1,9)
             if not isinstance(ws.cell(row=data_start,column=c),MergedCell)}
        for m in [str(m) for m in ws.merged_cells.ranges if m.min_row>=supply_row]:
            ws.unmerge_cells(m)
        ws.insert_rows(supply_row, add)
        for r in range(supply_row, supply_row+add):
            ws.merge_cells(f'A{r}:B{r}'); ws.row_dimensions[r].height=30.0
            for c,src in ref.items():
                dst=ws.cell(row=r,column=c)
                if not isinstance(dst,MergedCell): copy_style(src,dst)
        supply_row+=add; vat_row+=add; total_row+=add
        for rr in (supply_row,vat_row,total_row):
            ws.merge_cells(f'A{rr}:F{rr}'); ws.merge_cells(f'G{rr}:H{rr}')
            ws.row_dimensions[rr].height=42.75
        ws.row_dimensions[total_row+1].height=15.0
    for r in range(data_start, supply_row):
        for c in range(1,9):
            cell=ws.cell(row=r,column=c)
            if not isinstance(cell,MergedCell): cell.value=None
    for i,(d,item,spec,qty,price,amt) in enumerate(rows):
        r=data_start+i
        dc=ws.cell(row=r,column=1,value=d); dc.number_format='yyyy-mm-dd'
        ws.cell(row=r,column=3).value=item
        ws.cell(row=r,column=4).value=spec
        ws.cell(row=r,column=5).value=qty
        ws.cell(row=r,column=6).value=price
        ws.cell(row=r,column=7).value=amt
    ws.cell(row=supply_row,column=7).value=f'=SUM(G{data_start}:G{supply_row-1})'
    ws.cell(row=vat_row,   column=7).value='부가세별도'
    ws.cell(row=total_row, column=7).value=f'=SUM(G{supply_row}:G{vat_row})'

def save_to_folder(tmp, folder, fn):
    nfd=os.path.join(folder, unicodedata.normalize('NFD', fn)); nfc=os.path.join(folder, fn)
    for f in os.listdir(folder):
        if unicodedata.normalize('NFC',f)==fn:
            try: os.remove(os.path.join(folder,f))
            except OSError: pass
    wb=load_workbook(tmp); wb.save(nfd)
    try: os.rename(nfd,nfc)
    except OSError: pass

def app_root():
    p = os.path.abspath(__file__)
    for _ in range(8):
        p = os.path.dirname(p)
        if os.path.exists(os.path.join(p, 'server.js')):
            return p
    return ''

def template_score(path):
    name = unicodedata.normalize('NFC', os.path.basename(path))
    if name.startswith('~$') or not name.lower().endswith('.xlsx'):
        return -1
    named_for_haatz = '하츠' in name or 'haatz' in name.lower()
    score = 0
    if named_for_haatz:
        score += 45
    if '마감내역서' in name or 'template' in name.lower():
        score += 12
    try:
        size = os.path.getsize(path)
        score += min(size // 10000, 20)
        wb = load_workbook(path, read_only=True, data_only=False)
        if '판매현황' in wb.sheetnames:
            return -1
        sample = []
        for ws in wb.worksheets[:2]:
            for row in ws.iter_rows(max_row=min(ws.max_row, 14), max_col=min(ws.max_column, 8), values_only=True):
                sample.append(' '.join(str(x or '') for x in row))
        text = ' '.join(sample)
        if '거래  명  세  서' in text or '거래 명세서' in text:
            score += 35
        if '사업자등록번호' in text or '공급받는자' in text:
            score += 20
        if '일자' in text and '품목' in text and '공급가액' in text:
            score += 15
    except Exception:
        return -1
    if not named_for_haatz and ('하츠' not in text and 'haatz' not in text.lower()):
        return -1
    if score < 80:
        return -1
    return score

def template_search_dirs(template_arg, outdir, rawdir):
    dirs = []
    if template_arg and os.path.isdir(template_arg):
        dirs.append(template_arg)
    dirs.append(rawdir)
    root = app_root()
    if root:
        dirs.append(os.path.join(root, 'data', 'ai-skill-templates', 'haatz-ledger'))
        dirs.append(os.path.join(root, 'learning-data', '_무관_하츠'))
    seen, out = set(), []
    for d in dirs:
        if d and os.path.isdir(d):
            key = os.path.abspath(d).lower()
            if key not in seen:
                seen.add(key)
                out.append(d)
    return out

def pick_template(template_arg, outdir, rawdir):
    if template_arg and os.path.isfile(template_arg):
        return template_arg if template_score(template_arg) >= 0 else ''
    pools=[]
    for d in template_search_dirs(template_arg, outdir, rawdir):
        for f in os.listdir(d):
            full = os.path.join(d,f)
            if os.path.isfile(full) and template_score(full) >= 0:
                pools.append(full)
    return max(pools, key=lambda p: (template_score(p), os.path.getmtime(p))) if pools else ''

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--raw', required=True)
    ap.add_argument('--template', default='')
    ap.add_argument('--outdir', default='')
    args=ap.parse_args()
    raw=args.raw
    if not os.path.isfile(raw): print('[ERROR] 판매현황 없음:', raw); sys.exit(1)
    rawdir=os.path.dirname(os.path.abspath(raw)); outdir=args.outdir or rawdir
    os.makedirs(outdir, exist_ok=True)
    wb=load_workbook(raw)
    if '판매현황' not in wb.sheetnames: print("[ERROR] '판매현황' 시트 없음:", wb.sheetnames); sys.exit(3)
    ws=wb['판매현황']
    header=ws.cell(1,1).value or ''
    m=re.search(r'(\d{4})/(\d{2})/\d{2}\s*~', header)
    yy=m.group(1)[2:] if m else '00'; month_str=f'{m.group(2)}월' if m else '00월'
    m2=re.search(r'(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})', header)
    period=''
    if m2:
        d1=datetime.strptime(m2.group(1),'%Y/%m/%d'); d2=datetime.strptime(m2.group(2),'%Y/%m/%d')
        period=f'기  간: {d1.strftime("%Y-%m-%d")}부터 {d2.strftime("%Y-%m-%d")}까지'

    sites={}; warnings=[]
    for ri,row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row[0] is None: continue
        item=row[4]
        if not item or str(item).strip() in SKIP_ITEMS: continue
        site=str(row[2]).strip() if row[2] else '(현장미지정)'
        qty,price,amt=row[6],row[7],row[8]
        if qty and price and amt and abs(qty*price-amt)>1:
            warnings.append(f'행{ri} [{site}] {item} 검산오차')
        sites.setdefault(site, []).append((parse_date(row[0]), row[4], row[5], qty, price, amt))

    print('─'*60); print('판매현황:', raw); print('출력:', outdir)
    print('현장 수:', len(sites), '|', ', '.join(f'{k}({len(v)})' for k,v in sites.items()))
    if warnings: print('[WARN]', len(warnings),'건 검산오차')
    if not sites:
        print('[ERROR] 판매현황에서 처리할 하츠 현장/품목 데이터를 찾지 못했습니다.')
        sys.exit(5)

    tpl=pick_template(args.template, outdir, rawdir)
    if not tpl: print('[ERROR] 하츠 실제 전월 템플릿 없음 - 하츠 마감내역서 양식 xlsx를 먼저 등록해야 합니다.'); sys.exit(2)
    print('템플릿:', os.path.basename(tpl))

    made=[]
    for site, rows in sites.items():
        tmp_fd,tmp=tempfile.mkstemp(suffix='.xlsx', dir=os.environ.get('TEMP', tempfile.gettempdir()))
        os.close(tmp_fd); shutil.copy(tpl, tmp)
        try: os.chmod(tmp,0o644)
        except OSError: pass
        wbt=load_workbook(tmp)
        sheet='안전' if '안전' in wbt.sheetnames else wbt.sheetnames[0]
        ws_s=wbt[sheet]
        if period: ws_s['A4']=period
        ws_s['E4']=f'거래처: {site}'      # 하츠는 현장명을 E4 에 넣는다
        fill_sheet(ws_s, rows)
        wbt.save(tmp)
        safe=site.replace('/','_').strip() or site
        fn=f'{yy}년 {month_str}마감 하츠 마감내역서 - {safe}.xlsx'
        save_to_folder(tmp, outdir, fn)
        try: os.unlink(tmp)
        except OSError: pass
        made.append(fn); print(f'  -> {fn} ({len(rows)}건)')
    if not made:
        print('[ERROR] 생성된 하츠 마감 파일이 없습니다.')
        sys.exit(6)
    print(f'완료: {len(made)}개')

if __name__=='__main__':
    main()
