#!/usr/bin/env python3
"""E2E 업로드용 Excel 변환 스크립트
사용법: RAW_DATA_PATH와 OUTPUT_DIR 경로를 맞게 수정 후 실행
"""

import openpyxl, os, unicodedata, re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

RAW_DATA_PATH = ''      # ← 원본 판매현황 파일 경로
OUTPUT_DIR    = '/sessions/quirky-busy-gauss/mnt/09.마감자료 정리'

# ─────────────────────────────────────────────────────────────

def parse_date(s):
    """'2026/03/01 -244' → '2026-03-01'"""
    m = re.match(r'(\d{4}/\d{2}/\d{2})', str(s))
    return datetime.strptime(m.group(1), '%Y/%m/%d').strftime('%Y-%m-%d') if m else None

def make_e2e(raw_path, output_dir):
    wb_src = openpyxl.load_workbook(raw_path)
    ws_src = wb_src['판매현황']

    # 기간/거래처 추출
    header = ws_src.cell(1, 1).value or ''
    m_hdr = re.search(r'(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})', header)
    month_str = datetime.strptime(m_hdr.group(1), '%Y/%m/%d').strftime('%m') if m_hdr else '00'
    client = '대림컴퍼니'  # 기본값; 필요시 header에서 추출

    rows = []
    for r in ws_src.iter_rows(min_row=3, values_only=True):
        if not r[0]: continue
        date = parse_date(r[0])
        if not date or not r[4]: continue

        qty=r[6]; price=r[7]; amt=r[8]; vat=r[9]; total=r[10]

        # 수량×단가 ≠ 금액이면 소수점 단가 역산 보정
        if qty and price and amt and abs(qty * price - amt) > 0.01:
            orig = price
            price = round(amt / qty, 4)
            print(f'  단가 보정: {r[4]}  {orig} → {price}')

        rows.append({
            '일자':     date,
            '상품분류': '',
            '상품명':   r[4],
            '규격':     r[5] or '',
            '수량':     qty,
            '단가':     price,
            '금액':     amt,
            '세액':     vat,
            '합계금액': total,
            '비고':     r[11] or '',
            '약어':     '',
        })

    # 검증
    errs = [(r['상품명'], r['수량'], r['단가'], r['금액'])
            for r in rows
            if r['수량'] and r['단가'] and r['금액']
            and abs(r['수량'] * r['단가'] - r['금액']) > 0.01]
    if errs:
        print(f'⚠️ 검증 실패 {len(errs)}건:')
        for e in errs: print(f'   {e}')
    else:
        print(f'✅ 전체 {len(rows)}건 수량×단가=금액 검증 통과')

    # 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    HEADERS = ['일자','상품분류','상품명','규격','수량','단가','금액','세액','합계금액','비고','약어']
    COL_W   = [13, 10, 28, 18, 8, 14, 14, 12, 14, 12, 10]
    NUM_IDX = {5, 6, 7, 8, 9}   # E~I (1-based)

    thin = Side(style='thin')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더 (A열부터)
    for ci, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        c = ws.cell(1, ci, h)
        c.font      = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor='4472C4')
        c.border    = bdr
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20

    # 데이터
    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(HEADERS, 1):
            val = row[key]
            c   = ws.cell(ri, ci, val)
            c.font   = Font(name='맑은 고딕', size=10)
            c.border = bdr
            if ci in NUM_IDX:
                c.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(val, (int, float)):
                    is_dec = key == '단가' and isinstance(val, float) and val != int(val)
                    c.number_format = '#,##0.####' if is_dec else '#,##0'
            elif key == '일자':
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[ri].height = 17

    ws.freeze_panes = 'A2'

    # 저장 (한글 NFC)
    fname = f'{client}_{month_str}월_E2E업로드.xlsx'
    nfd   = os.path.join(output_dir, unicodedata.normalize('NFD', fname))
    nfc   = os.path.join(output_dir, fname)
    wb.save(nfd)
    if nfd != nfc:
        os.rename(nfd, nfc)
    print(f'✅ 저장: {fname}')
    return nfc

if __name__ == '__main__':
    if not RAW_DATA_PATH:
        print('RAW_DATA_PATH를 설정해주세요.')
    else:
        make_e2e(RAW_DATA_PATH, OUTPUT_DIR)
