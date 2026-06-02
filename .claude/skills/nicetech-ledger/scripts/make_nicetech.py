#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
나이스텍 마감내역서(거래명세서) 자동 생성 — ERP 에이전트용

퍼시스와 같은 "전월 템플릿 복제 → 데이터 채움" 방식. 단, 나이스텍은:
  - 시트 1개(안전). 잡자재 구분 없음. 전 품목 포함(매출할인만 제외)
  - 거래처(파일) 분리 = 판매현황 '프로젝트명' 컬럼. 비어있으면 비고 주소로 판별
      비고에 '코닝' → 한국코닝(주) / 그 외 빈칸 → 삼성SDI(기본)
  - E4 거래처명은 생성 파일의 거래처명으로 덮어씀
  - 부가세 행 = '부가세별도' 텍스트(수식 아님). 공급가액 G열만 사용

사용:
  python make_nicetech.py --raw <판매현황.xlsx> --template <전월 마감내역서.xlsx> --outdir <폴더>
  # template: '... 나이스텍 마감내역서 - {거래처}.xlsx' 형식. 거래처별로 매칭, 없으면 최신본 1개를 공용 틀로.
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import copy, shutil, os, glob, unicodedata, re, tempfile, argparse, sys
from datetime import datetime

SKIP_ITEMS = {'매출할인'}

def classify_vendor(proj, bigo):
    """프로젝트명 우선, 비어있으면 비고 주소로 거래처 판별."""
    p = str(proj).strip() if proj else ''
    if p:
        return p
    b = str(bigo) if bigo else ''
    if '코닝' in b: return '한국코닝(주)'
    return '삼성SDI'   # 비고로 못 가리면 삼성SDI 기본 (검증으로 확인)

def parse_date(s):
    m = re.match(r'(\d{4}/\d{2}/\d{2})', str(s))
    return datetime.strptime(m.group(1), '%Y/%m/%d') if m else None

def copy_style(src, dst):
    if src.has_style:
        dst.font=copy.copy(src.font); dst.border=copy.copy(src.border)
        dst.fill=copy.copy(src.fill); dst.number_format=src.number_format
        dst.alignment=copy.copy(src.alignment)

def find_summary_rows(ws):
    s=v=t=None
    for r in ws.iter_rows():
        for cell in r:
            if isinstance(cell, MergedCell): continue
            val=str(cell.value).strip() if cell.value else ''
            if val=='공급금액': s=cell.row
            elif val=='부가세': v=cell.row
            elif val=='총합계': t=cell.row
    return s,v,t

def fill_sheet(ws, rows, data_start=13):
    supply_row, vat_row, total_row = find_summary_rows(ws)
    if supply_row is None:
        raise RuntimeError("템플릿에서 '공급금액' 합계 행을 못 찾음 - 올바른 마감내역서 템플릿이 아님")
    available = supply_row - data_start
    needed = len(rows)
    if needed > available:
        add = needed - available
        ref = {c: ws.cell(row=data_start, column=c) for c in range(1,9)
               if not isinstance(ws.cell(row=data_start, column=c), MergedCell)}
        for m in [str(m) for m in ws.merged_cells.ranges if m.min_row >= supply_row]:
            ws.unmerge_cells(m)
        ws.insert_rows(supply_row, add)
        for r in range(supply_row, supply_row+add):
            ws.merge_cells(f'A{r}:B{r}'); ws.row_dimensions[r].height=30.0
            for c,src in ref.items():
                dst=ws.cell(row=r,column=c)
                if not isinstance(dst, MergedCell): copy_style(src,dst)
        supply_row+=add; vat_row+=add; total_row+=add
        for rr in (supply_row, vat_row, total_row):
            ws.merge_cells(f'A{rr}:F{rr}'); ws.merge_cells(f'G{rr}:H{rr}')
            ws.row_dimensions[rr].height=42.75
        ws.row_dimensions[total_row+1].height=15.0

    for r in range(data_start, supply_row):
        for c in range(1,9):
            cell=ws.cell(row=r,column=c)
            if not isinstance(cell, MergedCell): cell.value=None

    for i,(d,item,spec,qty,price,amt) in enumerate(rows):
        r=data_start+i
        dc=ws.cell(row=r,column=1,value=d); dc.number_format='yyyy-mm-dd'
        ws.cell(row=r,column=3).value=item
        ws.cell(row=r,column=4).value=spec
        ws.cell(row=r,column=5).value=qty
        ws.cell(row=r,column=6).value=price
        ws.cell(row=r,column=7).value=amt          # G=공급가액
        # H(비고)·부가세 수식 없음 — 나이스텍/하츠는 부가세별도

    ws.cell(row=supply_row, column=7).value=f'=SUM(G{data_start}:G{supply_row-1})'
    ws.cell(row=vat_row,    column=7).value='부가세별도'
    ws.cell(row=total_row,  column=7).value=f'=SUM(G{supply_row}:G{vat_row})'

def save_to_folder(tmp, folder, fn):
    nfd=os.path.join(folder, unicodedata.normalize('NFD', fn))
    nfc=os.path.join(folder, fn)
    for f in os.listdir(folder):
        if unicodedata.normalize('NFC', f)==fn:
            try: os.remove(os.path.join(folder,f))
            except OSError: pass
    wb=load_workbook(tmp); wb.save(nfd)
    try: os.rename(nfd, nfc)
    except OSError: pass

def app_root():
    p = os.path.abspath(__file__)
    for _ in range(8):
        p = os.path.dirname(p)
        if os.path.exists(os.path.join(p, 'server.js')):
            return p
    return ''

def template_score(path, vendor=''):
    """Prefer real previous closing templates over generated/simple scratch files."""
    name = unicodedata.normalize('NFC', os.path.basename(path))
    if name.startswith('~$') or not name.lower().endswith('.xlsx'):
        return -1
    score = 0
    normalized_name = re.sub(r'[^가-힣a-z0-9]', '', name.lower())
    if '나이스텍' in name or 'nicetech' in name.lower():
        score += 40
    if '마감내역서' in name or 'template' in name.lower():
        score += 12
    vendor_text = re.sub(r'\(주\)|㈜|주식회사|\s+', '', str(vendor).lower())
    vendor_text = vendor_text.replace('한국', '')
    vendor_key = re.sub(r'[^가-힣a-z0-9]', '', vendor_text)
    vendor_tokens = [vendor_key]
    vendor_tokens += [t for t in re.split(r'[^가-힣a-z0-9]+', vendor_text) if len(t) >= 2]
    for token in {t for t in vendor_tokens if len(t) >= 2}:
        if token in normalized_name:
            score += 35
    try:
        size = os.path.getsize(path)
        score += min(size // 10000, 20)
        wb = load_workbook(path, read_only=True, data_only=False)
        if '판매현황' in wb.sheetnames:
            return -1
        sample = []
        for ws in wb.worksheets[:2]:
            for row in ws.iter_rows(max_row=min(ws.max_row, 14), max_col=min(ws.max_column, 8), values_only=True):
                sample.append(' '.join(str(x or '') for x in row))
        text = ' '.join(sample)
        if '거래  명  세  서' in text or '거래 명세서' in text:
            score += 35
        if '사업자등록번호' in text or '공급받는자' in text:
            score += 20
        if '일자' in text and '품목' in text and '공급가액' in text:
            score += 15
    except Exception:
        return -1
    return score

def template_search_dirs(template_arg, outdir, rawdir):
    dirs = []
    if template_arg and os.path.isdir(template_arg):
        dirs.append(template_arg)
    dirs.append(rawdir)
    root = app_root()
    if root:
        dirs.append(os.path.join(root, 'data', 'ai-skill-templates', 'nicetech-ledger'))
        dirs.append(os.path.join(root, 'learning-data', '_무관_나이스텍'))
    seen, out = set(), []
    for d in dirs:
        if d and os.path.isdir(d):
            key = os.path.abspath(d).lower()
            if key not in seen:
                seen.add(key)
                out.append(d)
    return out

def pick_template(template_arg, vendor, outdir, rawdir):
    """거래처명이 파일명에 들어간 템플릿 우선, 없으면 최신 마감내역서 1개를 공용 틀로."""
    if template_arg and os.path.isfile(template_arg):
        return template_arg if template_score(template_arg, vendor) >= 0 else ''
    pools=[]
    for d in template_search_dirs(template_arg, outdir, rawdir):
        for f in os.listdir(d):
            full = os.path.join(d,f)
            if os.path.isfile(full) and template_score(full, vendor) >= 0:
                pools.append(full)
    if not pools: return ''
    # 거래처 키워드 매칭 (삼성/SDI, 코닝 등)
    return max(pools, key=lambda p: (template_score(p, vendor), os.path.getmtime(p)))

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--raw', required=True)
    ap.add_argument('--template', default='')
    ap.add_argument('--outdir', default='')
    args=ap.parse_args()

    raw=args.raw
    if not os.path.isfile(raw): print('[ERROR] 판매현황 없음:', raw); sys.exit(1)
    rawdir=os.path.dirname(os.path.abspath(raw))
    outdir=args.outdir or rawdir
    os.makedirs(outdir, exist_ok=True)

    wb=load_workbook(raw)
    if '판매현황' not in wb.sheetnames: print("[ERROR] '판매현황' 시트 없음:", wb.sheetnames); sys.exit(3)
    ws=wb['판매현황']

    # 헤더에서 월 추출
    header=ws.cell(1,1).value or ''
    m=re.search(r'(\d{4})/(\d{2})/\d{2}\s*~', header)
    month_str=f'{m.group(2)}월' if m else '00월'
    yy = m.group(1)[2:] if m else '00'

    vendors={}
    warnings=[]
    for ri,row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row[0] is None: continue
        item=row[4]
        if not item: continue
        if str(item).strip() in SKIP_ITEMS: continue
        vendor=classify_vendor(row[2], row[11] if len(row)>11 else '')
        qty,price,amt=row[6],row[7],row[8]
        if qty and price and amt and abs(qty*price-amt)>1:
            warnings.append(f'행{ri} [{vendor}] {item} 검산오차')
        vendors.setdefault(vendor, []).append((parse_date(row[0]), row[4], row[5], qty, price, amt))

    print('─'*60); print('판매현황:', raw); print('출력:', outdir)
    print('거래처 수:', len(vendors), '|', ', '.join(f'{k}({len(v)})' for k,v in vendors.items()))
    if warnings: print('[WARN]', len(warnings),'건 검산오차:', *warnings[:5], sep='\n  ')
    if not vendors:
        print('[ERROR] 판매현황에서 처리할 나이스텍 거래처/품목 데이터를 찾지 못했습니다.')
        sys.exit(5)

    template_by_vendor = {vendor: pick_template(args.template, vendor, outdir, rawdir) for vendor in vendors}
    made=[]
    for vendor, rows in vendors.items():
        tpl=template_by_vendor.get(vendor, '')
        if not tpl:
            print(f'[ERROR] [{vendor}] 템플릿 없음 - 전월 나이스텍 마감내역서 필요'); continue
        print(f'  템플릿: {os.path.basename(tpl)}')
        tmp_fd,tmp=tempfile.mkstemp(suffix='.xlsx', dir=os.environ.get('TEMP', tempfile.gettempdir()))
        os.close(tmp_fd); shutil.copy(tpl, tmp)
        try: os.chmod(tmp,0o644)
        except OSError: pass
        wbt=load_workbook(tmp)
        sheet = '안전' if '안전' in wbt.sheetnames else wbt.sheetnames[0]
        ws_s=wbt[sheet]
        # A4 기간과 E4 거래처를 현재 생성 파일 기준으로 갱신
        m2=re.search(r'(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})', header)
        if m2:
            d1=datetime.strptime(m2.group(1),'%Y/%m/%d'); d2=datetime.strptime(m2.group(2),'%Y/%m/%d')
            ws_s['A4']=f'기  간: {d1.strftime("%Y-%m-%d")}부터 {d2.strftime("%Y-%m-%d")}까지'
        ws_s['E4']=f'거래처: {vendor}'
        fill_sheet(ws_s, rows)
        wbt.save(tmp)
        # 파일명: 거래처명에서 안전한 이름
        vshort = vendor.replace('(주)','').replace('㈜','').replace('/','_').strip() or vendor
        fn=f'{yy}년 {month_str}마감 나이스텍 마감내역서 - {vshort}.xlsx'
        save_to_folder(tmp, outdir, fn)
        try: os.unlink(tmp)
        except OSError: pass
        made.append(fn); print(f'  -> {fn} ({len(rows)}건)')
    if not made:
        print('[ERROR] 생성된 나이스텍 마감 파일이 없습니다. 전월 나이스텍 마감내역서 템플릿을 함께 첨부하세요.')
        sys.exit(6)
    print(f'완료: {len(made)}개')

if __name__=='__main__':
    main()
