#!/usr/bin/env python3
"""Build a monthly Fursys/Persys ledger workbook from source XLSX files.

The script is intentionally deterministic so the AI does not re-invent the
ledger rules every time. It supports both monthly sales-status exports and
per-site statement files with 안전/잡자재 sheets.
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DETAIL_SHEET = "정리본"
SUMMARY_SHEET = "요약"
ISSUE_SHEET = "확인필요"
SUMMARY_LABELS = {"공급금액", "부가세", "총합계"}


DETAIL_HEADERS = [
    "원본파일",
    "원본시트",
    "원본행",
    "기간",
    "거래처명",
    "현장명",
    "거래일자",
    "전표번호",
    "구분",
    "품목코드",
    "품목",
    "규격",
    "수량",
    "단가",
    "공급가액",
    "부가세",
    "합계",
    "부가세기준",
    "검산결과",
    "비고",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def to_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def money_round(value: Any) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def calc_tax(supply: Any) -> int | None:
    if supply is None:
        return None
    return money_round(Decimal(str(supply)) * Decimal("0.1"))


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = clean_text(value)
    match = re.search(r"(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return text


def split_date_no(value: Any) -> tuple[str, str]:
    text = clean_text(value)
    if not text:
        return "", ""
    match = re.match(r"\s*(20\d{2}[-/.]\d{1,2}[-/.]\d{1,2})\s*-\s*(.+?)\s*$", text)
    if not match:
        return date_text(text), ""
    return date_text(match.group(1)), match.group(2)


def period_from_sheet(ws) -> str:
    for row in range(1, min(ws.max_row, 10) + 1):
        for col in range(1, ws.max_column + 1):
            text = clean_text(ws.cell(row, col).value)
            if "기간:" in text or "2026-04-01" in text or "2026/04/01" in text:
                return text
    return ""


def site_from_statement(ws, path: Path) -> str:
    for row in range(1, min(ws.max_row, 10) + 1):
        for col in range(1, ws.max_column + 1):
            text = clean_text(ws.cell(row, col).value)
            if text.startswith("거래처:"):
                return text.split(":", 1)[1].strip()
    stem = path.stem
    if stem.startswith("퍼시스-"):
        stem = stem[len("퍼시스-") :]
    if stem.endswith("-04월"):
        stem = stem[: -len("-04월")]
    return stem


def find_statement_header(ws) -> int | None:
    for row in range(1, min(ws.max_row, 35) + 1):
        if (
            clean_text(ws.cell(row, 1).value) == "일자"
            and clean_text(ws.cell(row, 3).value) == "품목"
            and clean_text(ws.cell(row, 7).value) == "공급가액"
        ):
            return row
    return None


def find_sales_header(ws) -> tuple[int, dict[str, int]] | None:
    wanted = {"일자-No.", "거래처명", "프로젝트명", "품목명", "수량", "단가", "공급가액", "부가세", "합계"}
    for row in range(1, min(ws.max_row, 20) + 1):
        values = [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if wanted.issubset(set(values)):
            return row, {name: idx + 1 for idx, name in enumerate(values) if name}
    return None


def make_record(
    *,
    path: Path,
    sheet: str,
    row_number: int,
    period: str,
    vendor: str,
    site: str,
    tx_date: str,
    slip_no: str,
    kind: str,
    item_code: str,
    item: str,
    spec: str,
    qty: Any,
    unit: Any,
    supply: Any,
    tax: Any,
    total: Any,
    tax_basis: str,
    note: str,
) -> dict[str, Any]:
    qty_n = to_number(qty)
    unit_n = to_number(unit)
    supply_n = to_number(supply)
    tax_n = to_number(tax)
    total_n = to_number(total)
    if tax_n is None and supply_n is not None:
        tax_n = calc_tax(supply_n)
        tax_basis = "10%계산"
    if total_n is None and supply_n is not None:
        total_n = (supply_n or 0) + (tax_n or 0)

    reasons: list[str] = []
    if not tx_date:
        reasons.append("날짜 없음")
    if not item:
        reasons.append("품목 없음")
    if supply_n is None:
        reasons.append("공급가액 없음")
    if qty_n is not None and unit_n is not None and supply_n is not None:
        expected = money_round(Decimal(str(qty_n)) * Decimal(str(unit_n)))
        if abs(expected - money_round(supply_n)) > 1:
            reasons.append(f"수량*단가 불일치: {expected} != {money_round(supply_n)}")
    if supply_n is not None and tax_n is not None and total_n is not None:
        expected_total = money_round(supply_n + tax_n)
        if abs(expected_total - money_round(total_n)) > 1:
            reasons.append(f"공급가액+부가세 불일치: {expected_total} != {money_round(total_n)}")

    return {
        "원본파일": path.name,
        "원본시트": sheet,
        "원본행": row_number,
        "기간": period,
        "거래처명": vendor,
        "현장명": site,
        "거래일자": tx_date,
        "전표번호": slip_no,
        "구분": kind,
        "품목코드": item_code,
        "품목": item,
        "규격": spec,
        "수량": qty_n,
        "단가": unit_n,
        "공급가액": money_round(supply_n) if supply_n is not None else None,
        "부가세": money_round(tax_n) if tax_n is not None else None,
        "합계": money_round(total_n) if total_n is not None else None,
        "부가세기준": tax_basis,
        "검산결과": "OK" if not reasons else "; ".join(reasons),
        "비고": note,
    }


def parse_sales_status(path: Path, wb) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        found = find_sales_header(ws)
        if not found:
            continue
        header_row, cols = found
        period = period_from_sheet(ws)
        for row_number in range(header_row + 1, ws.max_row + 1):
            row_values = [ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)]
            if not any(clean_text(v) for v in row_values):
                continue
            date_no = ws.cell(row_number, cols["일자-No."]).value
            tx_date, slip_no = split_date_no(date_no)
            vendor = clean_text(ws.cell(row_number, cols["거래처명"]).value)
            site = clean_text(ws.cell(row_number, cols["프로젝트명"]).value)
            item = clean_text(ws.cell(row_number, cols["품목명"]).value)
            qty = ws.cell(row_number, cols["수량"]).value
            unit = ws.cell(row_number, cols["단가"]).value
            supply = ws.cell(row_number, cols["공급가액"]).value
            if not item and to_number(qty) is None and to_number(unit) is None and to_number(supply) is None:
                continue
            records.append(
                make_record(
                    path=path,
                    sheet=ws.title,
                    row_number=row_number,
                    period=period,
                    vendor=vendor,
                    site=site,
                    tx_date=tx_date,
                    slip_no=slip_no,
                    kind=clean_text(ws.cell(row_number, cols.get("비고", 0)).value).split("/")[0].strip()
                    if cols.get("비고")
                    else "",
                    item_code=clean_text(ws.cell(row_number, cols.get("품목코드", 0)).value) if cols.get("품목코드") else "",
                    item=item,
                    spec=clean_text(ws.cell(row_number, cols.get("규격", 0)).value) if cols.get("규격") else "",
                    qty=qty,
                    unit=unit,
                    supply=supply,
                    tax=ws.cell(row_number, cols["부가세"]).value,
                    total=ws.cell(row_number, cols["합계"]).value,
                    tax_basis="원본",
                    note=clean_text(ws.cell(row_number, cols.get("비고", 0)).value) if cols.get("비고") else "",
                )
            )
    return records


def parse_statement_file(path: Path, wb) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        header_row = find_statement_header(ws)
        if header_row is None:
            continue
        site = site_from_statement(ws, path)
        period = period_from_sheet(ws)
        for row_number in range(header_row + 1, ws.max_row + 1):
            label = clean_text(ws.cell(row_number, 1).value)
            if label in SUMMARY_LABELS:
                break
            row_values = [ws.cell(row_number, col).value for col in range(1, 9)]
            if not any(clean_text(v) for v in row_values):
                continue
            records.append(
                make_record(
                    path=path,
                    sheet=ws.title,
                    row_number=row_number,
                    period=period,
                    vendor="(주)퍼시스",
                    site=site,
                    tx_date=date_text(ws.cell(row_number, 1).value),
                    slip_no="",
                    kind=ws.title,
                    item_code="",
                    item=clean_text(ws.cell(row_number, 3).value),
                    spec=clean_text(ws.cell(row_number, 4).value),
                    qty=ws.cell(row_number, 5).value,
                    unit=ws.cell(row_number, 6).value,
                    supply=ws.cell(row_number, 7).value,
                    tax=ws.cell(row_number, 8).value,
                    total=None,
                    tax_basis="원본" if to_number(ws.cell(row_number, 8).value) is not None else "10%계산",
                    note="",
                )
            )
    return records


def collect_source_files(inputs: list[str]) -> list[Path]:
    files: list[Path] = []
    for raw in inputs:
        path = Path(raw)
        if path.is_dir():
            files.extend(sorted(p for p in path.glob("*.xlsx") if not p.name.startswith("~$")))
        elif path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$"):
            files.append(path)
    seen: set[Path] = set()
    unique: list[Path] = []
    for path in files:
        resolved = path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            unique.append(path)
    return unique


def parse_sources(files: list[Path]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for path in files:
        wb = load_workbook(path, data_only=True, read_only=True)
        parsed = parse_statement_file(path, wb)
        if not parsed:
            parsed = parse_sales_status(path, wb)
        records.extend(parsed)
    return records


def write_summary(ws, records: list[dict[str, Any]], files: list[Path], issues: list[dict[str, Any]], source_label: str):
    site_summary = defaultdict(lambda: {"행수": 0, "수량합계": 0, "공급가액": 0, "부가세": 0, "합계": 0})
    kind_summary = defaultdict(lambda: {"행수": 0, "수량합계": 0, "공급가액": 0, "부가세": 0, "합계": 0})
    item_summary = defaultdict(lambda: {"행수": 0, "수량합계": 0, "공급가액": 0, "부가세": 0, "합계": 0})
    for rec in records:
        for key, bucket in [
            (rec["현장명"], site_summary),
            (rec["구분"], kind_summary),
            ((rec["품목"], rec["규격"]), item_summary),
        ]:
            data = bucket[key]
            data["행수"] += 1
            data["수량합계"] += rec["수량"] or 0
            data["공급가액"] += rec["공급가액"] or 0
            data["부가세"] += rec["부가세"] or 0
            data["합계"] += rec["합계"] or 0

    def append_section(title: str, headers: list[str], rows: list[list[Any]]):
        ws.append([title])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        ws.append([])

    append_section(
        "전체요약",
        ["항목", "값"],
        [
            ["원본", source_label],
            ["파일수", len(files)],
            ["정리행수", len(records)],
            ["공급가액", sum(r["공급가액"] or 0 for r in records)],
            ["부가세", sum(r["부가세"] or 0 for r in records)],
            ["합계", sum(r["합계"] or 0 for r in records)],
            ["확인필요", len(issues)],
        ],
    )
    append_section(
        "현장별 요약",
        ["현장명", "행수", "수량합계", "공급가액", "부가세", "합계"],
        [[k, v["행수"], v["수량합계"], v["공급가액"], v["부가세"], v["합계"]] for k, v in sorted(site_summary.items())],
    )
    append_section(
        "구분별 요약",
        ["구분", "행수", "수량합계", "공급가액", "부가세", "합계"],
        [[k, v["행수"], v["수량합계"], v["공급가액"], v["부가세"], v["합계"]] for k, v in sorted(kind_summary.items())],
    )
    item_rows = [[k[0], k[1], v["행수"], v["수량합계"], v["공급가액"], v["부가세"], v["합계"]] for k, v in item_summary.items()]
    item_rows.sort(key=lambda row: row[4], reverse=True)
    append_section("품목별 요약", ["품목", "규격", "행수", "수량합계", "공급가액", "부가세", "합계"], item_rows)


def style_workbook(wb: Workbook):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    issue_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        if ws.max_row:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = issue_fill if ws.title == ISSUE_SHEET else header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = 10
            for cell in ws[letter]:
                width = max(width, min(len("" if cell.value is None else str(cell.value)) + 2, 42))
            ws.column_dimensions[letter].width = width


def write_workbook(records: list[dict[str, Any]], files: list[Path], output: Path, source_label: str):
    wb = Workbook()
    detail = wb.active
    detail.title = DETAIL_SHEET
    detail.append(DETAIL_HEADERS)
    for rec in records:
        detail.append([rec.get(header) for header in DETAIL_HEADERS])

    issues = [rec for rec in records if rec.get("검산결과") != "OK"]
    summary = wb.create_sheet(SUMMARY_SHEET)
    write_summary(summary, records, files, issues, source_label)

    issue_ws = wb.create_sheet(ISSUE_SHEET)
    issue_headers = DETAIL_HEADERS + ["확인사유"]
    issue_ws.append(issue_headers)
    if issues:
        for rec in issues:
            issue_ws.append([rec.get(header) for header in DETAIL_HEADERS] + [rec.get("검산결과")])
    else:
        issue_ws.append(["확인필요 없음"] + [""] * (len(issue_headers) - 1))

    if records:
        table = Table(displayName="PersysDetail", ref=f"A1:{get_column_letter(detail.max_column)}{detail.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        detail.add_table(table)

    style_workbook(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return issues


def main() -> int:
    parser = argparse.ArgumentParser(description="Create a monthly Persys ledger workbook.")
    parser.add_argument("inputs", nargs="+", help="Source .xlsx files or folders containing source .xlsx files")
    parser.add_argument("--output", "-o", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    files = collect_source_files(args.inputs)
    if not files:
        raise SystemExit("No source .xlsx files found.")

    records = parse_sources(files)
    if not records:
        raise SystemExit("No Persys ledger rows were parsed from the source files.")

    output = Path(args.output)
    issues = write_workbook(records, files, output, source_label=", ".join(str(p) for p in files))
    print(f"output={output}")
    print(f"source_files={len(files)}")
    print(f"rows={len(records)}")
    print(f"issues={len(issues)}")
    print(f"supply={sum(r['공급가액'] or 0 for r in records)}")
    print(f"tax={sum(r['부가세'] or 0 for r in records)}")
    print(f"total={sum(r['합계'] or 0 for r in records)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
