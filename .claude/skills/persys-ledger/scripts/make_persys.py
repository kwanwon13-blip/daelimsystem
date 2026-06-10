#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
퍼시스 마감내역서(거래명세서) 자동 생성 — ERP 에이전트용 (CLI 인자판)

진짜 스킬 로직(전월 템플릿 복제 → 안전/잡자재 2시트 채움)을 그대로 쓰되,
경로 하드코딩 대신 CLI 인자를 받는다. ERP 에이전트가 첨부파일 경로로 호출 가능.

사용:
  python make_persys.py --raw <판매현황.xlsx> --template <전월 퍼시스-*.xlsx> --outdir <출력폴더>
  # template 생략 시 raw 파일과 같은 폴더(및 outdir)에서 퍼시스-*.xlsx 최신본 자동 탐색
  # outdir 생략 시 raw 파일과 같은 폴더

출력: 현장(프로젝트)별로 퍼시스-{현장명}-{MM월}.xlsx (안전/잡자재 2시트)
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import copy, shutil, os, glob, unicodedata, re, tempfile, argparse, sys, json
from datetime import datetime

# ── 품목 분류 규칙 (입력 실수 보정) ──────────────────────────────────
OVERRIDES = {
    '머리보호내피':              '안전',
    '장갑 3M(슈퍼그립200)':      '잡자재',
    '안전모 투구자동':            '안전',
    '갈대비A':                   '잡자재',
    '마대':                      '잡자재',
    '천막(텐텐지)1.8*100M 백색': '잡자재',
    '폴리베니아 3T':              '잡자재',
    '황색외곽쓰레받이':           '잡자재',
}
SKIP_ITEMS = {'매출할인'}

def get_cat(item, bigo):
    if item in SKIP_ITEMS: return None
    for k in OVERRIDES:                       # startswith — 괄호 변형도 매칭
        if str(item).startswith(k): return OVERRIDES[k]
    b = str(bigo) if bigo else ''
    if '안전' in b:   return '안전'
    if '잡자재' in b: return '잡자재'
    return '잡자재'                            # 비고 없으면 기본 잡자재

def parse_date(s):
    m = re.match(r'(\d{4}/\d{2}/\d{2})', str(s))
    return datetime.strptime(m.group(1), '%Y/%m/%d') if m else None

def copy_style(src, dst):
    if src.has_style:
        dst.font          = copy.copy(src.font)
        dst.border        = copy.copy(src.border)
        dst.fill          = copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment     = copy.copy(src.alignment)

def find_summary_rows(ws):
    s = v = t = None
    for r in ws.iter_rows():
        for cell in r:
            if isinstance(cell, MergedCell): continue
            val = str(cell.value).strip() if cell.value else ''
            if val == '공급금액': s = cell.row
            elif val == '부가세':  v = cell.row
            elif val == '총합계':  t = cell.row
    return s, v, t

def fill_sheet(ws, new_data, data_start=13):
    supply_row, vat_row, total_row = find_summary_rows(ws)
    if supply_row is None:
        raise RuntimeError("템플릿에서 '공급금액' 합계 행을 찾지 못함 - 올바른 퍼시스 템플릿이 아닙니다")
    available = supply_row - data_start
    needed    = len(new_data)

    if needed > available:
        rows_to_add = needed - available
        ref = {c: ws.cell(row=data_start, column=c)
               for c in range(1,9)
               if not isinstance(ws.cell(row=data_start, column=c), MergedCell)}
        merges_to_rm = [str(m) for m in ws.merged_cells.ranges if m.min_row >= supply_row]
        for m in merges_to_rm: ws.unmerge_cells(m)
        ws.insert_rows(supply_row, rows_to_add)
        for r in range(supply_row, supply_row + rows_to_add):
            ws.merge_cells(f'A{r}:B{r}')
            ws.row_dimensions[r].height = 30.0
            for c, src in ref.items():
                dst = ws.cell(row=r, column=c)
                if not isinstance(dst, MergedCell): copy_style(src, dst)
        supply_row += rows_to_add; vat_row += rows_to_add; total_row += rows_to_add
        ws.merge_cells(f'A{supply_row}:F{supply_row}'); ws.merge_cells(f'G{supply_row}:H{supply_row}')
        ws.merge_cells(f'A{vat_row}:F{vat_row}');      ws.merge_cells(f'G{vat_row}:H{vat_row}')
        ws.merge_cells(f'A{total_row}:F{total_row}');  ws.merge_cells(f'G{total_row}:H{total_row}')
        ws.row_dimensions[supply_row].height  = 42.75
        ws.row_dimensions[vat_row].height     = 42.75
        ws.row_dimensions[total_row].height   = 42.75
        ws.row_dimensions[total_row+1].height = 15.0

    # 데이터 영역 클리어
    for r in range(data_start, supply_row):
        for c in range(1,9):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None

    # 데이터 입력
    for i, (date_val, item, spec, qty, unit_price, supply_amt) in enumerate(new_data):
        r = data_start + i
        date_cell = ws.cell(row=r, column=1, value=date_val)
        date_cell.number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=3).value = item
        ws.cell(row=r, column=4).value = spec
        ws.cell(row=r, column=5).value = qty
        ws.cell(row=r, column=6).value = unit_price
        ws.cell(row=r, column=7).value = supply_amt
        ws.cell(row=r, column=8).value = f'=G{r}*0.1'

    ws.cell(row=supply_row, column=7).value = f'=SUM(G{data_start}:G{supply_row-1})'
    ws.cell(row=vat_row,    column=7).value = f'=SUM(H{data_start}:H{vat_row-1})'
    ws.cell(row=total_row,  column=7).value = f'=SUM(G{supply_row}:G{vat_row})'

def save_to_folder(tmp_path, folder, filename):
    dst_nfd = os.path.join(folder, unicodedata.normalize('NFD', filename))
    dst_nfc = os.path.join(folder, filename)
    for f in os.listdir(folder):
        if unicodedata.normalize('NFC', f) == filename:
            try: os.remove(os.path.join(folder, f))
            except OSError: pass
    wb = load_workbook(tmp_path)
    wb.save(dst_nfd)
    try: os.rename(dst_nfd, dst_nfc)
    except OSError: pass

def app_root():
    p = os.path.abspath(__file__)
    for _ in range(8):
        p = os.path.dirname(p)
        if os.path.exists(os.path.join(p, 'server.js')):
            return p
    return ''

def template_score(path):
    name = unicodedata.normalize('NFC', os.path.basename(path))
    if name.startswith('~$') or not name.lower().endswith('.xlsx'):
        return -1
    try:
        size = os.path.getsize(path)
        if size < 20000:
            return -1
        wb = load_workbook(path, read_only=False, data_only=False)
        if '판매현황' in wb.sheetnames:
            return -1
        if '안전' not in wb.sheetnames or '잡자재' not in wb.sheetnames:
            return -1
        score = min(size // 10000, 30)
        if '퍼시스' in name or 'fursys' in name.lower() or 'persys' in name.lower():
            score += 35
        for sheet_name in ('안전', '잡자재'):
            ws = wb[sheet_name]
            if find_summary_rows(ws)[0] is not None:
                score += 20
            sample = []
            for row in ws.iter_rows(max_row=min(ws.max_row, 14), max_col=min(ws.max_column, 8), values_only=True):
                sample.append(' '.join(str(x or '') for x in row))
            text = ' '.join(sample)
            if '거래  명  세  서' in text or '거래 명세서' in text:
                score += 20
            if '사업자등록번호' in text or '공급받는자' in text:
                score += 15
        return score if score >= 80 else -1
    except Exception:
        return -1

def template_search_dirs(template_arg):
    dirs = []
    if template_arg and os.path.isdir(template_arg):
        dirs.append(template_arg)
    root = app_root()
    if root:
        dirs.append(os.path.join(root, 'data', 'ai-skill-templates', 'persys-ledger'))
        dirs.append(os.path.join(root, 'learning-data', '_무관_퍼시스'))
    seen, out = set(), []
    for d in dirs:
        if d and os.path.isdir(d):
            key = os.path.abspath(d).lower()
            if key not in seen:
                seen.add(key)
                out.append(d)
    return out

def find_template(template_arg=''):
    if template_arg and os.path.isfile(template_arg):
        if template_score(template_arg) >= 0:
            return template_arg
    cands = []
    for d in template_search_dirs(template_arg):
        for f in os.listdir(d):
            full = os.path.join(d, f)
            if os.path.isfile(full) and template_score(full) >= 0:
                cands.append(full)
    return max(cands, key=lambda p: (template_score(p), os.path.getmtime(p))) if cands else ''

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--raw', required=True, help='판매현황 xlsx 경로')
    ap.add_argument('--template', default='', help='전월 퍼시스-*.xlsx 템플릿 (생략 시 자동 탐색)')
    ap.add_argument('--outdir', default='', help='출력 폴더 (생략 시 raw 와 동일 폴더)')
    args = ap.parse_args()

    raw = args.raw
    if not os.path.isfile(raw):
        print('[ERROR] 판매현황 파일 없음:', raw); sys.exit(1)
    outdir = args.outdir or os.path.dirname(os.path.abspath(raw))
    os.makedirs(outdir, exist_ok=True)
    template = find_template(args.template)
    if not template:
        print('[ERROR] 퍼시스 실제 전월 템플릿 없음 - data/ai-skill-templates/persys-ledger 폴더에 퍼시스 마감내역서 양식 xlsx를 등록해야 합니다.'); sys.exit(2)

    print('─'*60)
    print('판매현황 :', raw)
    print('템플릿   :', template)
    print('출력폴더 :', outdir)
    print('─'*60)

    wb_raw = load_workbook(raw)
    if '판매현황' not in wb_raw.sheetnames:
        print("[ERROR] '판매현황' 시트가 없습니다. 시트:", wb_raw.sheetnames); sys.exit(3)
    ws_raw = wb_raw['판매현황']

    projects = {}
    warnings = []
    recon_raw_rows = 0; recon_raw_total = 0.0
    recon_excluded_rows = 0; recon_excluded_total = 0.0
    recon_out_rows = 0; recon_out_total = 0.0
    for row_idx, row in enumerate(ws_raw.iter_rows(min_row=3, values_only=True), start=3):
        if row[0] is None: continue
        _supply = row[8] or 0
        recon_raw_rows += 1; recon_raw_total += _supply
        if row[4] in SKIP_ITEMS:
            recon_excluded_rows += 1; recon_excluded_total += _supply
        proj = row[2]
        if not proj: continue
        item = row[4]
        if not item: continue
        cat = get_cat(item, row[11])
        if cat is None: continue
        qty, unit_price, supply_amt = row[6], row[7], row[8]
        if not unit_price: warnings.append(f'행{row_idx} [{proj}] {item} — 단가 없음')
        if not supply_amt: warnings.append(f'행{row_idx} [{proj}] {item} — 공급가액 없음')
        if qty and unit_price and supply_amt and abs(qty*unit_price - supply_amt) > 1:
            warnings.append(f'행{row_idx} [{proj}] {item} — 수량×단가={qty*unit_price:.0f} ≠ 공급가액({supply_amt})')
        projects.setdefault(proj, {'안전': [], '잡자재': []})
        projects[proj][cat].append((parse_date(row[0]), row[4], row[5], qty, unit_price, supply_amt))
        recon_out_rows += 1; recon_out_total += (supply_amt or 0)

    if warnings:
        print(f'[WARN] 데이터 이상 {len(warnings)}건:')
        for w in warnings: print('  ', w)
    else:
        print('[OK] 단가/공급가액 검증 통과')

    print('[RECON] ' + json.dumps({
        'raw_rows': recon_raw_rows, 'raw_total': round(recon_raw_total),
        'excluded_rows': recon_excluded_rows, 'excluded_total': round(recon_excluded_total),
        'excluded_note': '매출할인', 'out_rows': recon_out_rows, 'out_total': round(recon_out_total),
    }, ensure_ascii=False))

    header = ws_raw.cell(row=1, column=1).value or ''
    m = re.search(r'(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})', header)
    if m:
        d1 = datetime.strptime(m.group(1), '%Y/%m/%d')
        d2 = datetime.strptime(m.group(2), '%Y/%m/%d')
        period_text = f'기  간: {d1.strftime("%Y-%m-%d")}부터 {d2.strftime("%Y-%m-%d")}까지'
        month_str   = d1.strftime('%m월')
    else:
        period_text = '기  간: 날짜를 찾을 수 없습니다'
        month_str   = '00월'

    print('기간:', period_text)
    print('현장 수:', len(projects))
    if not projects:
        print('[ERROR] 판매현황에서 처리할 퍼시스 현장/품목 데이터를 찾지 못했습니다.')
        sys.exit(5)

    made = []
    for proj, data in projects.items():
        s_data, m_data = data['안전'], data['잡자재']
        tmp_fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir=os.environ.get('TEMP', tempfile.gettempdir()))
        os.close(tmp_fd)
        shutil.copy(template, tmp)
        try: os.chmod(tmp, 0o644)
        except OSError: pass
        wb = load_workbook(tmp)
        if '안전' not in wb.sheetnames or '잡자재' not in wb.sheetnames:
            print("[ERROR] 템플릿에 '안전'/'잡자재' 시트가 없습니다:", wb.sheetnames); sys.exit(4)
        ws_s = wb['안전']; ws_s['A4'] = period_text; ws_s['E4'] = f'거래처:{proj}'
        fill_sheet(ws_s, s_data)
        fill_sheet(wb['잡자재'], m_data)
        wb.save(tmp)
        filename = f'퍼시스-{proj}-{month_str}.xlsx'
        save_to_folder(tmp, outdir, filename)
        try: os.unlink(tmp)
        except OSError: pass
        made.append(filename)
        print(f'  -> {filename}  (안전 {len(s_data)} / 잡자재 {len(m_data)})')

    if not made:
        print('[ERROR] 생성된 퍼시스 마감 파일이 없습니다.')
        sys.exit(6)
    print(f'완료: {len(made)}개 파일 생성')

if __name__ == '__main__':
    main()
