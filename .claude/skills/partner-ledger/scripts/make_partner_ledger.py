#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create partner monthly closing statement workbooks from an eCount sales workbook."""

from __future__ import annotations

import argparse
import copy
import os
import re
import shutil
import sys
import tempfile
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import MergedCell


# ⚠️ SSOT: lib/agent-runtime.js 의 PARTNER_LEDGER_VENDORS 와 동일하게 유지할 것.
#    (JS↔Python 런타임 분리로 import 불가 — 거래처 추가/변경 시 양쪽 동시 수정)
SUPPORTED = [
    "한신공영",
    "요진건설",
    "홍지이앤씨",
    "삼진비티",
    "선두종합기술",
    "익스테리어앤",
    "한국지오텍",
    "금광스틸",
]


def clean_text(value) -> str:
    return str(value or "").strip()


def sanitize_filename(name: str) -> str:
    name = unicodedata.normalize("NFC", clean_text(name))
    name = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    return name or "마감내역서"


def vendor_alias(raw_vendor: str) -> str:
    s = clean_text(raw_vendor)
    checks = [
        ("한신", "한신공영"),
        ("요진", "요진건설"),
        ("홍지", "홍지이앤씨"),
        ("삼진", "삼진비티"),
        ("선두", "선두종합기술"),
        ("익스테리어", "익스테리어앤"),
        ("지오텍", "한국지오텍"),
        ("금광", "금광스틸"),
    ]
    for needle, alias in checks:
        if needle in s:
            return alias
    s = re.sub(r"주식회사|\(주\)|㈜", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_date(value):
    if isinstance(value, datetime):
        return value
    s = clean_text(value)
    m = re.search(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", s)
    if not m:
        return None
    return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))


def parse_month_from_header(value) -> str:
    s = clean_text(value)
    m = re.search(r"(\d{4})[./-](\d{1,2})[./-]\d{1,2}", s)
    if m:
        return f"{int(m.group(2))}월"
    return f"{datetime.now().month}월"


def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def cell_value(row, header_map, name):
    idx = header_map.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def find_raw_sheet_and_headers(wb):
    required = {"일자-No.", "거래처명", "프로젝트명", "품목명", "수량", "단가", "공급가액"}
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 20) + 1):
            values = [clean_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 40) + 1)]
            found = {v for v in values if v}
            if len(required.intersection(found)) >= 6:
                header_map = {}
                for i, value in enumerate(values):
                    if value and value not in header_map:
                        header_map[value] = i
                return ws, r, header_map
    raise RuntimeError("판매현황 헤더를 찾지 못했습니다. 원본 파일인지 확인해주세요.")


def read_raw_rows(raw_path: str):
    wb = load_workbook(raw_path, data_only=True)
    ws, header_row, header_map = find_raw_sheet_and_headers(wb)
    month_label = parse_month_from_header(ws.cell(1, 1).value)
    rows = []
    warnings = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        item = clean_text(cell_value(row, header_map, "품목명"))
        if not item:
            continue
        vendor_raw = clean_text(cell_value(row, header_map, "거래처명"))
        project = clean_text(cell_value(row, header_map, "프로젝트명"))
        if not project:
            continue
        qty = cell_value(row, header_map, "수량")
        unit_price = cell_value(row, header_map, "단가")
        supply = cell_value(row, header_map, "공급가액")
        try:
            if qty is not None and unit_price is not None and supply is not None:
                calc = float(qty) * float(unit_price)
                if abs(calc - float(supply)) > 1:
                    warnings.append(f"row {row_idx}: {project} {item} 공급가액 확인 필요 ({calc:.0f} != {supply})")
        except Exception:
            warnings.append(f"row {row_idx}: {project} {item} 수량/단가/공급가액 숫자 확인 필요")
        rows.append({
            "row_idx": row_idx,
            "date": parse_date(cell_value(row, header_map, "일자-No.")),
            "vendor_raw": vendor_raw,
            "vendor": vendor_alias(vendor_raw),
            "project": project,
            "item": item,
            "spec": cell_value(row, header_map, "규격"),
            "qty": qty,
            "unit_price": unit_price,
            "supply": supply,
        })
    return rows, month_label, warnings


def app_root() -> str:
    p = os.path.abspath(__file__)
    for _ in range(8):
        p = os.path.dirname(p)
        if os.path.exists(os.path.join(p, "server.js")):
            return p
    return ""


def is_partner_template(path: str) -> bool:
    try:
        if not path or not os.path.isfile(path):
            return False
        name = os.path.basename(path)
        if name.startswith("~$") or not name.lower().endswith(".xlsx"):
            return False
        wb = load_workbook(path, read_only=True, data_only=False)
        ws = wb.worksheets[0]
        values = [clean_text(ws.cell(1, c).value) for c in range(1, min(ws.max_column, 20) + 1)]
        found = {v for v in values if v}
        required = {"일자-No.", "품목명", "수량", "단가", "공급가액"}
        raw_markers = {"거래처명", "프로젝트명", "품목코드"}
        return len(required.intersection(found)) >= 5 and not raw_markers.intersection(found) and ws.max_column <= 10
    except Exception:
        return False


def template_search_dirs(template_arg: str):
    roots = []
    if template_arg:
        p = Path(template_arg)
        if p.is_dir():
            roots.append(str(p))
    root = app_root()
    if root:
        roots.append(os.path.join(root, "data", "ai-skill-templates", "partner-ledger"))
        roots.append(os.path.join(root, "learning-data", "_무관_협력사"))
    seen, out = set(), []
    for root in roots:
        if root and os.path.isdir(root):
            key = os.path.abspath(root).lower()
            if key not in seen:
                seen.add(key)
                out.append(root)
    return out


def list_templates(template_arg: str, raw_dir: str, outdir: str, raw_path: str):
    explicit = []
    candidates = []
    if template_arg:
        p = Path(template_arg)
        if p.is_file() and p.suffix.lower() == ".xlsx" and is_partner_template(str(p)):
            explicit.append(str(p))
    for root in template_search_dirs(template_arg):
        if not root or not os.path.isdir(root):
            continue
        for name in os.listdir(root):
            full = os.path.join(root, name)
            if is_partner_template(full):
                candidates.append(full)
    unique = []
    seen = set()
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    for path in explicit + candidates:
        key = os.path.abspath(path).lower()
        if key not in seen:
            seen.add(key)
            unique.append(path)
    return unique


def choose_template(templates, vendor: str):
    if not templates:
        return ""
    vendor_norm = sanitize_filename(vendor).replace(" ", "")
    for path in templates:
        name_norm = sanitize_filename(os.path.basename(path)).replace(" ", "")
        if vendor_norm and vendor_norm in name_norm:
            return path
    for alias in SUPPORTED:
        if alias == vendor:
            continue
        # Prefer an unmatched single template over using a different named vendor template.
    if len(templates) == 1:
        return templates[0]
    return templates[0]


def remove_existing_equivalent(folder: str, filename: str):
    target = unicodedata.normalize("NFC", filename)
    for name in os.listdir(folder):
        if unicodedata.normalize("NFC", name) == target:
            try:
                os.remove(os.path.join(folder, name))
            except OSError:
                pass


def prepare_sheet(ws, row_count: int):
    if row_count <= 0:
        return
    max_col = max(6, ws.max_column)
    if ws.max_row < 2:
        ws.append(["", "", "", "", "", ""])
    template_height = ws.row_dimensions[2].height
    template_styles = {}
    for c in range(1, max_col + 1):
        cell = ws.cell(2, c)
        if not isinstance(cell, MergedCell):
            template_styles[c] = copy.copy(cell)
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    for c in range(1, max_col + 1):
        cell = ws.cell(2, c)
        if not isinstance(cell, MergedCell):
            cell.value = None
    if row_count > 1:
        ws.insert_rows(3, row_count - 1)
    for r in range(2, 2 + row_count):
        if template_height is not None:
            ws.row_dimensions[r].height = template_height
        for c, src in template_styles.items():
            dst = ws.cell(r, c)
            if not isinstance(dst, MergedCell):
                copy_cell_style(src, dst)


def populate_workbook(template_path: str, output_path: str, rows):
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    shutil.copy(template_path, tmp_path)
    wb = load_workbook(tmp_path)
    ws = wb["판매현황"] if "판매현황" in wb.sheetnames else wb.worksheets[0]
    rows = sorted(rows, key=lambda x: (x["date"] or datetime.min, x["row_idx"]))
    prepare_sheet(ws, len(rows))
    previous_date = None
    for i, item in enumerate(rows, start=2):
        date_value = item["date"]
        ws.cell(i, 1).value = date_value if date_value and date_value != previous_date else None
        if date_value:
            ws.cell(i, 1).number_format = "yyyy-mm-dd"
            previous_date = date_value
        ws.cell(i, 2).value = item["item"]
        ws.cell(i, 3).value = item["spec"]
        ws.cell(i, 4).value = item["qty"]
        ws.cell(i, 5).value = item["unit_price"]
        ws.cell(i, 6).value = f"=D{i}*E{i}"
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    remove_existing_equivalent(os.path.dirname(output_path), os.path.basename(output_path))
    wb.save(output_path)
    try:
        os.unlink(tmp_path)
    except OSError:
        pass


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", required=True, help="eCount 판매현황 xlsx")
    ap.add_argument("--template", default="", help="template xlsx file or template directory")
    ap.add_argument("--outdir", default="", help="output directory")
    args = ap.parse_args()

    raw = os.path.abspath(args.raw)
    if not os.path.isfile(raw):
        print("[ERROR] 원본 판매현황 파일이 없습니다:", raw)
        sys.exit(1)
    raw_dir = os.path.dirname(raw)
    outdir = os.path.abspath(args.outdir or raw_dir)
    os.makedirs(outdir, exist_ok=True)

    rows, month_label, warnings = read_raw_rows(raw)
    supported_rows = [r for r in rows if r["vendor"] in SUPPORTED]
    if supported_rows:
        rows = supported_rows
    if not rows:
        print("[ERROR] 처리할 판매현황 행이 없습니다.")
        sys.exit(2)

    templates = list_templates(args.template, raw_dir, outdir, raw)
    if not templates:
        print("[ERROR] 마감내역서 템플릿 파일이 필요합니다. 템플릿 xlsx를 함께 첨부하거나 등록해주세요.")
        sys.exit(3)

    groups = defaultdict(list)
    for row in rows:
        groups[(row["vendor"], row["project"])].append(row)

    print("─" * 60)
    print("원본:", raw)
    print("출력:", outdir)
    print("템플릿 후보:", len(templates))
    print("그룹 수:", len(groups))
    if warnings:
        print(f"[WARN] 금액 확인 필요 {len(warnings)}건")
        for warning in warnings[:20]:
            print("  -", warning)

    made = []
    for (vendor, project), items in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        template = choose_template(templates, vendor)
        if not template:
            print(f"[ERROR] [{vendor}] 템플릿을 찾지 못했습니다.")
            continue
        filename = sanitize_filename(f"{vendor} {project} {month_label} 마감내역서.xlsx")
        output_path = os.path.join(outdir, filename)
        populate_workbook(template, output_path, items)
        made.append(filename)
        print(f"  -> {filename} ({len(items)}건, 템플릿: {os.path.basename(template)})")

    if not made:
        print("[ERROR] 생성된 마감내역서가 없습니다.")
        sys.exit(4)
    print(f"완료: {len(made)}개 파일 생성")


if __name__ == "__main__":
    main()
